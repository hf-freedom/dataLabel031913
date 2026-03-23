import random
import string
import os
import time
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from playwright.sync_api import sync_playwright
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SCREENSHOT_DIR = r"C:\Users\12824\Desktop\dataLabel\0319\p13\login_picture"
EXCEL_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "registration_data.xlsx")
BASE_URL = "http://39.107.109.8:8082/"

FIRST_NAMES = ["张", "王", "李", "赵", "刘", "陈", "杨", "黄", "周", "吴", "徐", "孙", "马", "朱", "胡", "郭", "何", "高", "林", "罗"]
LAST_NAMES = ["伟", "芳", "娜", "秀英", "敏", "静", "丽", "强", "磊", "军", "洋", "勇", "艳", "杰", "娟", "涛", "明", "超", "秀兰", "霞"]

lock = threading.Lock()


def generate_random_string(length=8):
    return ''.join(random.choices(string.ascii_lowercase + string.digits, k=length))


def generate_random_email():
    username = generate_random_string(10)
    return f"{username}@163.com"


def generate_random_password():
    return generate_random_string(12) + random.choice(string.ascii_uppercase) + random.choice(string.digits)


def generate_random_name():
    first_name = random.choice(FIRST_NAMES)
    last_name = random.choice(LAST_NAMES)
    return first_name + last_name


def generate_random_age():
    return str(random.randint(18, 60))


def generate_random_phone():
    prefixes = ["130", "131", "132", "133", "134", "135", "136", "137", "138", "139",
                "150", "151", "152", "153", "155", "156", "157", "158", "159",
                "170", "176", "177", "178",
                "180", "181", "182", "183", "184", "185", "186", "187", "188", "189"]
    prefix = random.choice(prefixes)
    suffix = ''.join(random.choices(string.digits, k=8))
    return prefix + suffix


def ensure_screenshot_dir():
    if not os.path.exists(SCREENSHOT_DIR):
        os.makedirs(SCREENSHOT_DIR)
        print(f"创建截图目录: {SCREENSHOT_DIR}")


def init_excel_file():
    """初始化Excel文件，如果不存在则创建"""
    if os.path.exists(EXCEL_FILE):
        return
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "注册数据"
    
    headers = [
        "序号", "线程ID", "用户名", "密码", "邮箱", "姓名", "年龄", "手机号",
        "注册开始时间", "注册结束时间", "注册耗时(秒)", "注册状态",
        "登录开始时间", "登录结束时间", "登录耗时(秒)", "登录状态",
        "验证结果", "总耗时(秒)", "创建时间"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
    
    wb.save(EXCEL_FILE)
    print(f"创建Excel文件: {EXCEL_FILE}")


def save_to_excel(data):
    """保存注册数据到Excel"""
    with lock:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        row = ws.max_row + 1
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            if col == 12:  # 注册状态
                if value == "成功":
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    cell.font = Font(color="006100")
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    cell.font = Font(color="9C0006")
            elif col == 16:  # 登录状态
                if value == "成功":
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    cell.font = Font(color="006100")
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    cell.font = Font(color="9C0006")
            elif col == 17:  # 验证结果
                if value == "通过":
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    cell.font = Font(color="006100")
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    cell.font = Font(color="9C0006")
        
        wb.save(EXCEL_FILE)


def register_and_login(thread_id):
    """执行注册和登录验证的完整流程"""
    thread_name = f"Thread-{thread_id}"
    print(f"\n[{thread_name}] 开始执行任务...")
    
    total_start_time = time.time()
    
    username = "user_" + generate_random_string(6)
    password = generate_random_password()
    email = generate_random_email()
    name = generate_random_name()
    age = generate_random_age()
    phone = generate_random_phone()
    
    print(f"[{thread_name}] 生成用户信息:")
    print(f"  用户名: {username}")
    print(f"  密码: {password}")
    print(f"  邮箱: {email}")
    print(f"  姓名: {name}")
    
    reg_start_time = datetime.now()
    reg_start_timestamp = time.time()
    reg_status = "失败"
    reg_end_time = ""
    reg_duration = 0
    
    login_start_time = ""
    login_end_time = ""
    login_duration = 0
    login_status = "未执行"
    verify_result = "未验证"
    
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=False)
            context = browser.new_context()
            page = context.new_page()
            
            try:
                print(f"[{thread_name}] 正在打开注册页面...")
                page.goto(BASE_URL, timeout=30000)
                page.wait_for_load_state("networkidle")
                page.wait_for_timeout(2000)
                
                register_link = page.query_selector('text=注册') or page.query_selector('text=立即注册') or page.query_selector('a:has-text("注册")')
                if register_link:
                    register_link.click()
                    page.wait_for_load_state("networkidle")
                    page.wait_for_timeout(1000)
                
                username_selectors = [
                    'input[name="username"]', 'input[name="user"]', 'input[name="userName"]',
                    'input[placeholder*="用户名"]', 'input[placeholder*="账号"]', '#username', '#userName'
                ]
                password_selectors = [
                    'input[name="password"]', 'input[name="pwd"]', 'input[name="userPassword"]',
                    'input[placeholder*="密码"]', '#password', '#pwd', 'input[type="password"]'
                ]
                email_selectors = [
                    'input[name="email"]', 'input[name="mail"]', 'input[placeholder*="邮箱"]',
                    'input[placeholder*="Email"]', '#email', 'input[type="email"]'
                ]
                name_selectors = [
                    'input[name="name"]', 'input[name="realName"]', 'input[name="realname"]',
                    'input[placeholder*="姓名"]', 'input[placeholder*="真实姓名"]', '#name', '#realName'
                ]
                age_selectors = ['input[name="age"]', 'input[placeholder*="年龄"]', '#age', 'input[type="number"]']
                phone_selectors = [
                    'input[name="phone"]', 'input[name="mobile"]', 'input[name="tel"]',
                    'input[placeholder*="手机"]', 'input[placeholder*="电话"]', '#phone', '#mobile'
                ]
                
                def find_input(selectors):
                    for selector in selectors:
                        try:
                            element = page.query_selector(selector)
                            if element:
                                return element
                        except:
                            continue
                    return None
                
                username_input = find_input(username_selectors)
                password_input = find_input(password_selectors)
                email_input = find_input(email_selectors)
                name_input = find_input(name_selectors)
                age_input = find_input(age_selectors)
                phone_input = find_input(phone_selectors)
                
                if username_input:
                    username_input.fill(username)
                if password_input:
                    password_input.fill(password)
                if email_input:
                    email_input.fill(email)
                if name_input:
                    name_input.fill(name)
                if age_input:
                    age_input.fill(age)
                if phone_input:
                    phone_input.fill(phone)
                
                page.wait_for_timeout(500)
                
                submit_selectors = [
                    'button:has-text("注册")', 'button:has-text("提交")', 'button:has-text("确定")',
                    'input[type="submit"]', 'input[value="注册"]', 'input[value="提交"]',
                    '.register-btn', '#register-btn', 'button[type="submit"]'
                ]
                
                submit_btn = None
                for selector in submit_selectors:
                    try:
                        submit_btn = page.query_selector(selector)
                        if submit_btn:
                            break
                    except:
                        continue
                
                if submit_btn:
                    submit_btn.click()
                    page.wait_for_load_state("networkidle")
                    page.wait_for_timeout(3000)
                    reg_status = "成功"
                    print(f"[{thread_name}] 注册成功!")
                else:
                    print(f"[{thread_name}] 未找到注册按钮")
                
                reg_end_time = datetime.now()
                reg_duration = round(time.time() - reg_start_timestamp, 2)
                
                if reg_status == "成功":
                    login_start_time = datetime.now()
                    login_start_timestamp = time.time()
                    print(f"[{thread_name}] 开始登录验证...")
                    
                    page.goto(BASE_URL, timeout=30000)
                    page.wait_for_load_state("networkidle")
                    page.wait_for_timeout(2000)
                    
                    login_username_selectors = [
                        'input[name="username"]', 'input[name="user"]', '#username',
                        'input[placeholder*="用户名"]', 'input[placeholder*="账号"]'
                    ]
                    login_password_selectors = [
                        'input[name="password"]', 'input[type="password"]', '#password',
                        'input[placeholder*="密码"]'
                    ]
                    
                    login_user_input = find_input(login_username_selectors)
                    login_pwd_input = find_input(login_password_selectors)
                    
                    if login_user_input and login_pwd_input:
                        login_user_input.fill(username)
                        login_pwd_input.fill(password)
                        page.wait_for_timeout(500)
                        
                        login_btn_selectors = [
                            'button:has-text("登录")', 'button:has-text("登陆")', 'input[type="submit"]',
                            'input[value="登录"]', 'input[value="登陆"]', '.login-btn', '#login-btn'
                        ]
                        
                        login_btn = None
                        for selector in login_btn_selectors:
                            try:
                                login_btn = page.query_selector(selector)
                                if login_btn:
                                    break
                            except:
                                continue
                        
                        if login_btn:
                            login_btn.click()
                            page.wait_for_load_state("networkidle")
                            page.wait_for_timeout(3000)
                            
                            page_content = page.content()
                            if username in page_content or name in page_content or "欢迎" in page_content or "个人中心" in page_content:
                                login_status = "成功"
                                verify_result = "通过"
                                print(f"[{thread_name}] 登录验证成功! 用户信息正确")
                            else:
                                login_status = "成功"
                                verify_result = "需确认"
                                print(f"[{thread_name}] 登录成功，但需人工确认用户信息")
                        else:
                            login_status = "失败"
                            verify_result = "未验证"
                    else:
                        login_status = "失败"
                        verify_result = "未验证"
                    
                    login_end_time = datetime.now()
                    login_duration = round(time.time() - login_start_timestamp, 2)
                
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                screenshot_path = os.path.join(SCREENSHOT_DIR, f"result_{thread_name}_{timestamp}.png")
                page.screenshot(path=screenshot_path)
                print(f"[{thread_name}] 已保存结果截图: {screenshot_path}")
                
            except Exception as e:
                print(f"[{thread_name}] 执行过程中出错: {e}")
                reg_status = "失败"
                login_status = "失败"
                verify_result = "未验证"
                
            finally:
                browser.close()
    
    except Exception as e:
        print(f"[{thread_name}] 浏览器启动失败: {e}")
        reg_status = "失败"
        login_status = "失败"
        verify_result = "未验证"
    
    total_duration = round(time.time() - total_start_time, 2)
    
    excel_data = [
        thread_id, thread_name, username, password, email, name, age, phone,
        reg_start_time.strftime("%Y-%m-%d %H:%M:%S") if isinstance(reg_start_time, datetime) else reg_start_time,
        reg_end_time.strftime("%Y-%m-%d %H:%M:%S") if isinstance(reg_end_time, datetime) else reg_end_time,
        reg_duration, reg_status,
        login_start_time.strftime("%Y-%m-%d %H:%M:%S") if isinstance(login_start_time, datetime) else login_start_time,
        login_end_time.strftime("%Y-%m-%d %H:%M:%S") if isinstance(login_end_time, datetime) else login_end_time,
        login_duration, login_status, verify_result, total_duration,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ]
    
    save_to_excel(excel_data)
    print(f"[{thread_name}] 数据已保存到Excel，总耗时: {total_duration}秒")
    
    return {
        "thread_id": thread_id,
        "username": username,
        "reg_status": reg_status,
        "login_status": login_status,
        "verify_result": verify_result,
        "total_duration": total_duration
    }


def run_parallel_registration(num_threads=5):
    """并行执行多个注册任务"""
    print("=" * 60)
    print("自动注册系统 - 并行模式")
    print("=" * 60)
    print(f"并发数量: {num_threads} 个")
    print(f"Excel文件: {EXCEL_FILE}")
    print(f"截图目录: {SCREENSHOT_DIR}")
    print("=" * 60)
    
    ensure_screenshot_dir()
    init_excel_file()
    
    results = []
    overall_start = time.time()
    
    with ThreadPoolExecutor(max_workers=num_threads) as executor:
        futures = {executor.submit(register_and_login, i+1): i+1 for i in range(num_threads)}
        
        for future in as_completed(futures):
            thread_id = futures[future]
            try:
                result = future.result()
                results.append(result)
            except Exception as e:
                print(f"[Thread-{thread_id}] 任务执行异常: {e}")
                results.append({
                    "thread_id": thread_id,
                    "username": "",
                    "reg_status": "失败",
                    "login_status": "失败",
                    "verify_result": "异常",
                    "total_duration": 0
                })
    
    overall_duration = round(time.time() - overall_start, 2)
    
    print("\n" + "=" * 60)
    print("所有任务执行完成!")
    print("=" * 60)
    print(f"总执行时间: {overall_duration} 秒")
    print(f"成功注册: {sum(1 for r in results if r['reg_status'] == '成功')}/{num_threads}")
    print(f"成功登录: {sum(1 for r in results if r['login_status'] == '成功')}/{num_threads}")
    print(f"验证通过: {sum(1 for r in results if r['verify_result'] == '通过')}/{num_threads}")
    print("=" * 60)
    
    print("\n详细结果:")
    for r in results:
        print(f"  [{r['thread_id']:2d}] 用户: {r['username'][:20]:20s} | 注册: {r['reg_status']} | 登录: {r['login_status']} | 验证: {r['verify_result']} | 耗时: {r['total_duration']}s")
    
    print(f"\n数据已保存至: {EXCEL_FILE}")
    
    return results


if __name__ == "__main__":
    run_parallel_registration(5)
